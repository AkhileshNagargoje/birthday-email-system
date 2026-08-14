"""Reading and writing the student CSV from the dashboard.

Edits preserve whatever columns the file already had (a Google Forms export
has a Timestamp column, for instance), so nothing is lost on save.
"""

import csv
import io
import re
import shutil
from datetime import datetime
from pathlib import Path

from . import config, gitstore, students as roster

CANONICAL = ["Full Name", "Email Address", "Date of Birth"]

# Path inside the repository, used only when the GitHub backend is on.
REPO_PATH = "data/students.csv"


def _repo_path():
    """Where the roster lives in the repo, derived from SOURCE_PATH."""
    try:
        return config.SOURCE_PATH.relative_to(config.ROOT).as_posix()
    except ValueError:
        return REPO_PATH


def _sync_down():
    """Pulls the roster from GitHub onto local disk.

    The action buttons shell out to src.main, which reads the file, so the
    local copy has to exist even when GitHub is the real store.
    """
    if not gitstore.enabled():
        return
    text, _ = gitstore.fetch(_repo_path())
    if text is None:
        return
    config.SOURCE_PATH.parent.mkdir(parents=True, exist_ok=True)
    config.SOURCE_PATH.write_text(text, encoding="utf-8")


def sync():
    """Public face of _sync_down, for callers outside this module."""
    _sync_down()
    # The sent log too, so "last sent" reflects what the workflow recorded.
    if gitstore.enabled():
        text, _ = gitstore.fetch("logs/sent_log.csv")
        if text is not None:
            config.SENT_LOG.parent.mkdir(parents=True, exist_ok=True)
            config.SENT_LOG.write_text(text, encoding="utf-8")


def _read_raw():
    _sync_down()
    path = config.SOURCE_PATH
    if not path.exists():
        return [], list(CANONICAL)
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        rows = list(reader)
        fields = list(reader.fieldnames or CANONICAL)
    return rows, fields


def _columns(fields):
    return (
        roster._find_column(fields, roster.NAME_KEYS) or CANONICAL[0],
        roster._find_column(fields, roster.EMAIL_KEYS) or CANONICAL[1],
        roster._find_column(fields, roster.DOB_KEYS) or CANONICAL[2],
    )


def _write(rows, fields, what="Update students"):
    path = config.SOURCE_PATH
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        # One rolling backup, so a bad edit is always undoable.
        shutil.copy2(path, path.with_suffix(path.suffix + ".bak"))
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

    if gitstore.enabled():
        # Commit it back, otherwise the edit dies with the next restart.
        gitstore.put(_repo_path(),
                     path.read_text(encoding="utf-8"),
                     f"{what} (via dashboard)")


def list_students():
    """Every row, valid or not, so the dashboard can show what needs fixing."""
    rows, fields = _read_raw()
    name_c, email_c, dob_c = _columns(fields)

    out = []
    for i, row in enumerate(rows):
        name = str(row.get(name_c) or "").strip()
        email = str(row.get(email_c) or "").strip()
        dob_raw = row.get(dob_c)
        dob = roster.parse_dob(dob_raw)

        problem = None
        if not name:
            problem = "no name"
        elif not email:
            problem = "no email"
        elif not roster.EMAIL_RE.match(email.lower()):
            problem = "email looks wrong"
        elif dob is None:
            problem = f"date not understood: {dob_raw!r}"

        out.append({
            "index": i,
            "name": name,
            "email": email,
            "dob": dob.isoformat() if dob else str(dob_raw or ""),
            "dob_display": dob.strftime("%d %b %Y") if dob else str(dob_raw or ""),
            "problem": problem,
        })
    return out


def add_student(name, email, dob):
    rows, fields = _read_raw()
    name_c, email_c, dob_c = _columns(fields)

    email = email.strip().lower()
    if any(str(r.get(email_c) or "").strip().lower() == email for r in rows):
        raise ValueError(f"{email} is already on the list.")

    row = {f: "" for f in fields}
    row[name_c] = name.strip()
    row[email_c] = email
    row[dob_c] = dob.strip()

    ts_c = roster._find_column(fields, roster.TIMESTAMP_KEYS)
    if ts_c:
        row[ts_c] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    rows.append(row)
    _write(rows, fields)


def update_student(index, name, email, dob):
    rows, fields = _read_raw()
    if not 0 <= index < len(rows):
        raise ValueError("That student is no longer in the list - reload the page.")
    name_c, email_c, dob_c = _columns(fields)

    email = email.strip().lower()
    for i, r in enumerate(rows):
        if i != index and str(r.get(email_c) or "").strip().lower() == email:
            raise ValueError(f"{email} is already used by another student.")

    rows[index][name_c] = name.strip()
    rows[index][email_c] = email
    rows[index][dob_c] = dob.strip()
    _write(rows, fields)


def parse_upload(filename, data):
    """Turns an uploaded file into rows of {name, email, dob}.

    Column headers are matched loosely - the same rules the daily job uses -
    so a Google Forms export works untouched, whatever the questions were
    called.
    """
    suffix = Path(filename).suffix.lower()

    if suffix in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError(
                "Reading .xlsx needs openpyxl: pip install openpyxl"
            ) from exc
        ws = load_workbook(io.BytesIO(data), data_only=True, read_only=True).active
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [str(h).strip() if h is not None else "" for h in next(rows)]
        except StopIteration:
            raise ValueError("That spreadsheet is empty.")
        records = [dict(zip(headers, row)) for row in rows]

    elif suffix in (".csv", ".tsv", ".txt"):
        text = data.decode("utf-8-sig", errors="replace")
        if not text.strip():
            raise ValueError("That file is empty.")
        # Sniff the separator so tab- and semicolon-separated exports work too.
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        records = list(csv.DictReader(io.StringIO(text), dialect=dialect))
        headers = list(records[0].keys()) if records else []

    else:
        raise ValueError(
            f"Cannot read {suffix or 'that file'}. Save it as .csv or .xlsx first "
            f"- from Excel or Google Sheets, File → Download → CSV."
        )

    if not records:
        raise ValueError("No rows found in that file.")

    name_c = roster._find_column(headers, roster.NAME_KEYS)
    email_c = roster._find_column(headers, roster.EMAIL_KEYS)
    dob_c = roster._find_column(headers, roster.DOB_KEYS)

    missing = [label for label, col in
               (("name", name_c), ("email", email_c), ("date of birth", dob_c))
               if col is None]
    if missing:
        raise ValueError(
            "Could not find a column for " + ", ".join(missing) +
            ". Headers found: " + ", ".join(str(h) for h in headers if h)
        )

    return [{"name": r.get(name_c), "email": r.get(email_c), "dob": r.get(dob_c)}
            for r in records]


def import_students(incoming):
    """Merges rows into the roster. Existing people are updated by email
    rather than duplicated; unreadable rows are reported, never guessed at.

    Returns (added, updated, skipped) where skipped is [(row_number, reason)].
    """
    rows, fields = _read_raw()
    name_c, email_c, dob_c = _columns(fields)

    by_email = {}
    for i, row in enumerate(rows):
        key = str(row.get(email_c) or "").strip().lower()
        if key:
            by_email[key] = i

    added = updated = 0
    skipped = []
    seen = set()

    for n, raw in enumerate(incoming, start=2):   # 2 = first row under the header
        name = str(raw.get("name") or "").strip()
        email = str(raw.get("email") or "").strip().lower()
        dob_raw = raw.get("dob")

        if not any((name, email, dob_raw)):
            continue                              # blank line, not an error

        if not name:
            skipped.append((n, "no name"))
            continue
        if not email:
            skipped.append((n, "no email address"))
            continue
        if not roster.EMAIL_RE.match(email):
            skipped.append((n, f"invalid email: {email}"))
            continue
        if email in seen:
            skipped.append((n, f"appears twice in the file: {email}"))
            continue

        parsed = roster.parse_dob(dob_raw)
        if parsed is None:
            skipped.append((n, f"unreadable date of birth: {dob_raw!r}"))
            continue

        seen.add(email)
        iso = parsed.isoformat()

        if email in by_email:
            row = rows[by_email[email]]
            if (str(row.get(name_c) or "").strip() != name
                    or roster.parse_dob(row.get(dob_c)) != parsed):
                row[name_c] = name
                row[dob_c] = iso
                updated += 1
        else:
            row = {f: "" for f in fields}
            row[name_c] = name
            row[email_c] = email
            row[dob_c] = iso
            ts_c = roster._find_column(fields, roster.TIMESTAMP_KEYS)
            if ts_c:
                row[ts_c] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            rows.append(row)
            by_email[email] = len(rows) - 1
            added += 1

    if added or updated:
        _write(rows, fields, f"Import: {added} added, {updated} updated")

    return added, updated, skipped


def delete_student(index):
    rows, fields = _read_raw()
    if not 0 <= index < len(rows):
        raise ValueError("That student is no longer in the list - reload the page.")
    removed = rows.pop(index)
    _write(rows, fields)
    name_c, _, _ = _columns(fields)
    return str(removed.get(name_c) or "that student")


# ---------------------------------------------------------------------------
# .env editing
# ---------------------------------------------------------------------------

EDITABLE = [
    "EMAIL_USER", "EMAIL_PASS", "EMAIL_FROM_NAME", "REPLY_TO",
    "SMTP_HOST", "SMTP_PORT", "TEST_EMAIL", "REPORT_EMAIL", "REPORT_ALWAYS",
    "BCC_EMAIL", "SEND_DELAY_SECONDS", "POSTER_MODE",
]


def read_settings():
    path = config.ROOT / ".env"
    values = {key: "" for key in EDITABLE}
    if not path.exists():
        return values
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key = key.strip()
        if key in values:
            values[key] = value.strip()
    return values


def write_settings(updates):
    """Rewrites .env in place, keeping comments and unknown keys intact."""
    path = config.ROOT / ".env"
    if not path.exists():
        example = config.ROOT / ".env.example"
        path.write_text(example.read_text(encoding="utf-8") if example.exists() else "",
                        encoding="utf-8")

    lines = path.read_text(encoding="utf-8").splitlines()
    remaining = dict(updates)

    for i, line in enumerate(lines):
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key = stripped.split("=", 1)[0].strip()
        if key in remaining:
            lines[i] = f"{key}={remaining.pop(key)}"

    for key, value in remaining.items():
        lines.append(f"{key}={value}")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


SAFE_VALUE = re.compile(r"^[^\r\n]*$")


def validate_setting(key, value):
    if not SAFE_VALUE.match(value):
        raise ValueError(f"{key} cannot contain a line break.")
    if key == "SMTP_PORT" and value and not value.isdigit():
        raise ValueError("SMTP_PORT must be a number.")
    return value.strip()
