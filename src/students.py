"""Reads the student roster from CSV / XLSX / Google Sheet and cleans it up.

Google Form headers are whatever the questions were called, so columns are
matched loosely ("Date of Birth", "DOB", "Birth date" all work) rather than
requiring exact names.
"""

import csv
import re
from dataclasses import dataclass
from datetime import date, datetime

from . import config

# Column header keywords, in priority order. First hit wins.
NAME_KEYS = ["full name", "student name", "your name", "name"]
EMAIL_KEYS = ["email address", "email", "e-mail", "mail id", "mail"]
DOB_KEYS = ["date of birth", "birth date", "birthday", "birthdate", "dob"]
TIMESTAMP_KEYS = ["timestamp", "submitted at", "date submitted"]

# Tried in order. Day-first comes before month-first because Indian college
# forms overwhelmingly use DD/MM/YYYY.
DATE_FORMATS = [
    "%Y-%m-%d",
    "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y",
    "%m/%d/%Y", "%m-%d-%Y",
    "%d/%m/%y", "%d-%m-%y",
    "%d %B %Y", "%d %b %Y",
    "%B %d, %Y", "%b %d, %Y",
    "%B %d %Y", "%b %d %Y",
    "%Y/%m/%d",
]

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[A-Za-z]{2,}$")


@dataclass
class Student:
    name: str
    email: str
    dob: date

    @property
    def first_name(self):
        return self.name.split()[0] if self.name.split() else self.name

    def age_turning(self, on_day):
        return on_day.year - self.dob.year


@dataclass
class BadRow:
    row_number: int
    reason: str
    raw: dict


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------

def load_rows():
    """Returns a list of dicts straight from the source, headers untouched."""
    if config.SOURCE_TYPE == "csv":
        return _load_csv(config.SOURCE_PATH)
    if config.SOURCE_TYPE == "xlsx":
        return _load_xlsx(config.SOURCE_PATH)
    if config.SOURCE_TYPE == "gsheet":
        return _load_gsheet()
    raise ValueError(
        f"Unknown SOURCE_TYPE {config.SOURCE_TYPE!r}. Use csv, xlsx or gsheet."
    )


def _load_csv(path):
    if not path.exists():
        raise FileNotFoundError(
            f"Student file not found: {path}\n"
            f"Copy data/students.example.csv to {path.name} and fill it in, "
            f"or point SOURCE_PATH at your file."
        )
    # utf-8-sig strips the BOM Excel and Google Sheets like to add.
    with open(path, newline="", encoding="utf-8-sig") as fh:
        return list(csv.DictReader(fh))


def _load_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("SOURCE_TYPE=xlsx needs openpyxl: pip install openpyxl") from exc

    if not path.exists():
        raise FileNotFoundError(f"Student file not found: {path}")

    ws = load_workbook(path, data_only=True).active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    return [dict(zip(headers, row)) for row in rows]


def _load_gsheet():
    try:
        import gspread
    except ImportError as exc:
        raise ImportError(
            "SOURCE_TYPE=gsheet needs: pip install gspread google-auth"
        ) from exc

    if not config.GSHEET_ID:
        raise ValueError("Set GSHEET_ID in .env (the long id from the sheet URL).")
    if not config.GOOGLE_CREDENTIALS_FILE.exists():
        raise FileNotFoundError(
            f"Service account key not found: {config.GOOGLE_CREDENTIALS_FILE}"
        )

    gc = gspread.service_account(filename=str(config.GOOGLE_CREDENTIALS_FILE))
    book = gc.open_by_key(config.GSHEET_ID)
    try:
        ws = book.worksheet(config.GSHEET_WORKSHEET)
    except Exception:
        ws = book.sheet1
    return ws.get_all_records()


# ---------------------------------------------------------------------------
# Cleaning
# ---------------------------------------------------------------------------

def _find_column(headers, keywords):
    lowered = {h: str(h).strip().lower() for h in headers if h}
    for key in keywords:
        for original, low in lowered.items():
            if low == key:
                return original
    for key in keywords:
        for original, low in lowered.items():
            if key in low:
                return original
    return None


def parse_dob(value):
    """Best-effort date parsing. Returns a date, or None if unreadable."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in DATE_FORMATS:
        try:
            parsed = datetime.strptime(text, fmt).date()
        except ValueError:
            continue
        # "%y" turns 04 into 2004, but strptime maps 69-99 to 1969-1999,
        # which is what we want for birthdates anyway.
        if parsed.year < 1900 or parsed.year > date.today().year:
            continue
        return parsed
    return None


def parse_rows(rows):
    """Turns raw rows into (students, bad_rows). Deduplicates by email."""
    if not rows:
        return [], []

    headers = list(rows[0].keys())
    name_col = _find_column(headers, NAME_KEYS)
    email_col = _find_column(headers, EMAIL_KEYS)
    dob_col = _find_column(headers, DOB_KEYS)
    ts_col = _find_column(headers, TIMESTAMP_KEYS)

    missing = [
        label
        for label, col in (("name", name_col), ("email", email_col), ("date of birth", dob_col))
        if col is None
    ]
    if missing:
        raise ValueError(
            "Could not find a column for: "
            + ", ".join(missing)
            + f"\nHeaders seen: {headers}\n"
            "Rename the column, or add the header text to the *_KEYS lists in src/students.py."
        )

    by_email = {}
    bad = []

    for i, row in enumerate(rows, start=2):  # start=2 → matches spreadsheet row numbers
        name = str(row.get(name_col) or "").strip()
        email = str(row.get(email_col) or "").strip().lower()
        dob_raw = row.get(dob_col)

        if not name:
            bad.append(BadRow(i, "missing name", row))
            continue
        if not email:
            bad.append(BadRow(i, "missing email", row))
            continue
        if not EMAIL_RE.match(email):
            bad.append(BadRow(i, f"invalid email: {email}", row))
            continue

        dob = parse_dob(dob_raw)
        if dob is None:
            bad.append(BadRow(i, f"unreadable date of birth: {dob_raw!r}", row))
            continue

        student = Student(name=_titlecase(name), email=email, dob=dob)
        # Later submissions win, so a student who re-submits to fix a typo is honoured.
        by_email[email] = (row.get(ts_col) if ts_col else i, student)

    students = [s for _, s in by_email.values()]
    students.sort(key=lambda s: s.name.lower())
    return students, bad


def _titlecase(name):
    """'PRAJWAL shete' -> 'Prajwal Shete', but leaves MixedCase names alone."""
    if name.isupper() or name.islower():
        return " ".join(part.capitalize() for part in name.split())
    return " ".join(name.split())


def load_students():
    return parse_rows(load_rows())


# ---------------------------------------------------------------------------
# Birthday matching
# ---------------------------------------------------------------------------

def _is_leap(year):
    return year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)


def birthdays_on(students, day):
    """Students whose birthday falls on `day`.

    Feb 29 birthdays are wished on Feb 28 in non-leap years, so they never
    get skipped for three years running.
    """
    matches = [s for s in students if (s.dob.month, s.dob.day) == (day.month, day.day)]

    if day.month == 2 and day.day == 28 and not _is_leap(day.year):
        matches += [s for s in students if (s.dob.month, s.dob.day) == (2, 29)]

    return matches
